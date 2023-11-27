# -*- coding: utf-8 -*-
import argparse
import hashlib
import os
import sys
import time
import logging
from itertools import groupby

from openpyxl import load_workbook

logging.basicConfig(format='%(asctime)s - %(filename)s[line:%(lineno)d] - %(levelname)s: %(message)s',
                    filename="init.log", level=logging.DEBUG)

script_version = "20220302-1"

ENV = os.getenv("TOOLS_ENV")
DEV_WORKSPACE = "/Users/kane/Documents/gingkoo/imas/generator/"
WORKSPACE = "./" if ENV is None else DEV_WORKSPACE
db_type = "oracle"

tpl_dict = {}
basic_info = {}  # 基本信息
branch_list = []  # 机构
report_org_list = []  # 上报行
role_list = []  # 角色
business_line_list = []  # 业务线
user_list = []  # 用户
function_list = []  # 角色功能权限
object_privilege_list = []  # 业务线对象权限
report_list = []
report_field_list = {}  # 报表字段信息
sys_param_config = {}  # 系统初始化参数配置信息

sql_list = []

curr_date_str = time.strftime("%Y%m%d", time.localtime())
curr_datetime_str = time.strftime("%Y%m%d%H%M%S", time.localtime())

admin_role_id = "31aa52b7fdb34749969bce5673abab7d"

role_func_rel_tpl = "insert into GP_BM_ROLE_FUNC_REL(ID, ROLE_ID, FUNCID, NEXT_ACTION, DATA_STATUS) \
select {}, '{}', {}, '99','04' from GP_BM_FUNCTION where FUNCID not like 'BTN@%'"

role_func_rel_tpl2 = "insert into GP_BM_ROLE_FUNC_REL(ID, ROLE_ID, FUNCID, NEXT_ACTION, DATA_STATUS) \
select {}, '{}', FUNCID, '99','04' from GP_BM_FUNCTION where FUNCID like 'BTN@%'"


def upper_camel_case(source):
    return "".join(map(lambda x: x.capitalize(), source.split("_")))


def load_config(tpl_path):
    wb = load_workbook(tpl_path)
    load_tpl_dict(wb)
    load_basic_info(wb)
    if basic_info["tpl_version"] != script_version:
        raise Exception("脚本版本号[{}]与模板版本号[{}]不匹配".format(script_version, basic_info["tpl_version"]))
    load_branches(wb)
    load_report_orgs(wb)
    load_roles(wb)
    load_business_lines(wb)
    load_users(wb)
    load_functions(wb)
    load_reports(wb)
    load_report_fields(wb)
    load_object_privilege(wb)
    load_sys_param(wb)


def gen_sql():
    # 更新配置表中的org_id
    sql_list.append("delete from GP_BM_ROLE_FUNC_REL where ROLE_ID = '{}'".format(admin_role_id))
    if db_type == "mysql":
        sql_list.append(role_func_rel_tpl.format("uuid()", admin_role_id, "CONCAT('PAGE@', FUNCID)"))
        sql_list.append(role_func_rel_tpl.format("uuid()", admin_role_id, "FUNCID"))
        sql_list.append(role_func_rel_tpl2.format("uuid()", admin_role_id))
    elif db_type == "oracle":
        sql_list.append(role_func_rel_tpl.format("sys_guid()", admin_role_id, "'PAGE@' || FUNCID"))
        sql_list.append(role_func_rel_tpl.format("sys_guid()", admin_role_id, "FUNCID"))
        sql_list.append(role_func_rel_tpl2.format("sys_guid()", admin_role_id))
    elif db_type == "sybase":
        sql_list.append(role_func_rel_tpl.format("newid()", admin_role_id, "'PAGE@' || FUNCID"))
        sql_list.append(role_func_rel_tpl.format("newid()", admin_role_id, "FUNCID"))
        sql_list.append(role_func_rel_tpl2.format("newid()", admin_role_id))
    elif db_type == "sqlserver":
        sql_list.append(role_func_rel_tpl.format("newId()", admin_role_id, "'PAGE@' + FUNCID"))
        sql_list.append(role_func_rel_tpl.format("newId()", admin_role_id, "FUNCID"))
        sql_list.append(role_func_rel_tpl2.format("newId()", admin_role_id))
    sql_list.append(
        "update GP_BM_ID_FILEDATA set DEPART_ID = '{}' where DEPART_ID is not null".format(basic_info["bank_org_id"]))
    sql_list.append(
        "update GP_BM_ROLE_INFO set ORG_ID = '{}' where ROLE_NAME = '管理岗_admin'".format(basic_info["bank_org_id"]))
    # 初始化数据生成
    gp_bm_sys_stat()
    gp_bm_branch()
    # bm_rpt_org()
    # bm_todo_task_cfg()
    gp_bm_role_info()
    gp_bm_role_func_rel()
    gp_bm_business_line()
    gp_bm_business_line_role()
    gp_bm_res_tpl_inst()
    gp_bm_res_tpl_field()
    gp_bm_tlr_info()
    gp_bm_tlr_role_rel()
    gp_bm_tlr_org_rel()
    # gp_bm_sys_param()


def gp_bm_sys_stat():
    # 数据清理
    sql_list.append("delete from GP_BM_SYS_STAT")
    data = {
        "DATA_ID": md5_str(str(basic_info["bank_org_id"]) + str(basic_info["module_name"])),
        "CORP_ID": basic_info["bank_code"],
        "ORG_ID": basic_info["bank_org_id"],
        "SYSTEM_NAME": basic_info["module_name"],
        "SYS_DATE": curr_date_str,
        "LAST_WORK_DATE": curr_date_str,
        "WORK_DATE": curr_date_str,
        "BH_DATE": curr_date_str,
        "STATUS": "0",
    }
    add_default_cols(data, True)
    sql_list.append(dict_to_sqls("GP_BM_SYS_STAT", data))


def gp_bm_branch():
    # 数据清理
    sql_list.append("delete from GP_BM_BRANCH")
    for branch_info in branch_list:
        data = {
            "DATA_ID": md5_str(str(branch_info["id"])),
            "DATA_DATE": curr_date_str,
            "CORP_ID": basic_info["bank_code"],
            "ORG_ID": basic_info["bank_org_id"],
            "BRCODE": branch_info["id"],
            "BRNO": branch_info["id"],
            "BRNAME": branch_info["name"],
            "BRATTR": branch_info["type"],
            "BRCLASS": branch_info["type"],
            "BLN_BRANCH_BRCODE": '01',
            "BLN_UP_BRCODE": branch_info["pid"],
            "STATUS": "1",
            "ST": "4",
            "IS_LOCK": "0",
            "IS_DEL": "F",
            "GPMS_NEXTACTION": "21",
        }
        add_default_cols(data)
        sql_list.append(dict_to_sqls("GP_BM_BRANCH", data))


def bm_rpt_org():
    sql_list.append("delete from " + basic_info["sys_id"] + "_BM_RPT_ORG")
    for report_org_info in report_org_list:
        data = {
            "DATA_ID": md5_str(str(report_org_info["id"]) + str(report_org_info["module_id"])),
            "CORP_ID": basic_info["bank_code"],
            "ORG_ID": basic_info["bank_org_id"],
            "NBJGH": report_org_info["id"],
            "JRJGDM": report_org_info["uni_credit_code"],
            "JGMC": report_org_info["name"],
            "IS_REPORT": report_org_info["is_report"],
            "P_NBJGH": report_org_info["p_nbjgh"],
            "BANK_ORG_CODE": report_org_info["bank_org_code"],
            "MODULE_ID": report_org_info["module_id"],
            "DATA_DATE": curr_date_str,
        }
        sql_list.append(dict_to_sqls(basic_info["sys_id"] + "_BM_RPT_ORG", data))


def bm_todo_task_cfg():
    sql_list.append("delete from " + basic_info["sys_id"] + "_BM_TODO_TASK_CFG")
    report_list_sorted = sorted(report_list, key=lambda report_info: str(report_info["id"]).split("_")[0])
    report_map = groupby(report_list_sorted, key=lambda report_info: str(report_info["id"]).split("_")[0])
    report_org_map = {str(report_org_info["module_id"]): report_org_info for report_org_info in report_org_list}
    for moduleId, report_list_module in report_map:
        report_org_info_module = report_org_map[moduleId]
        for report_info_module in list(report_list_module):
            data_id = md5_str(str(report_org_info_module["id"]) + str(report_info_module["id"]))
            data = {
                "DATA_ID": data_id,
                "DATA_DATE": curr_date_str,
                "CORP_ID": basic_info["bank_code"],
                "ORG_ID": basic_info["bank_org_id"],
                "TASK_MODULE": str(report_info_module["id"]).split("_")[0],
                "TASK_ID": data_id,
                "TASK_TYPE": "0",
                "REPORT_ORG_ID": report_org_info_module["id"],
                "REPORT_CODE": report_info_module["id"],
                "REMARKS": report_info_module["name"],
                "DATA_STATUS": "Y",
            }
            sql_list.append(dict_to_sqls(basic_info["sys_id"] + "_BM_TODO_TASK_CFG", data))
            print(report_info_module)


def gp_bm_role_info():
    sql_list.append(f"delete from GP_BM_ROLE_INFO WHERE ROLE_ID!='{admin_role_id}'")
    for role_info in role_list:
        if role_info["id"] == admin_role_id:
            sql_list.append(f"delete from GP_BM_ROLE_INFO WHERE ROLE_ID='{admin_role_id}'")
        data = {
            "DATA_ID": role_info["id"],
            "CORP_ID": basic_info["bank_code"],
            "ROLE_ID": role_info["id"],
            "ROLE_NAME": role_info["name"],
            "ROLE_TYPE": "1",
            "STATUS": "1",
            "GPMS_NEXTACTION": "00",
        }
        add_default_cols(data, True)
        sql_list.append(dict_to_sqls("GP_BM_ROLE_INFO", data))


def gp_bm_role_func_rel():
    sql_list.append(f"delete from GP_BM_ROLE_FUNC_REL where ROLE_ID!='{admin_role_id}'")
    for role_info in role_list:
        if role_info["id"] == admin_role_id:
            continue
        for function_info in role_info["functions"]:
            if function_info["type"] in ["菜单", "导航"]:
                # 菜单权限
                menu_id = function_info["id"]
                role_rel_info = {
                    "ID": md5_str(str(role_info["id"]) + str(menu_id)),
                    "ROLE_ID": role_info["id"],
                    "FUNCID": menu_id,
                    "CORP_ID": basic_info["bank_code"],
                    "ORG_ID": basic_info["bank_org_id"],
                    "GROUP_ID": "HZ"
                }
                sql_list.append(dict_to_sqls("GP_BM_ROLE_FUNC_REL", role_rel_info))
                # 页面权限
                page_id = "PAGE@" + function_info["id"]
                role_rel_info = {
                    "ID": md5_str(str(role_info["id"]) + str(page_id)),
                    "ROLE_ID": role_info["id"],
                    "FUNCID": page_id,
                    "CORP_ID": basic_info["bank_code"],
                    "ORG_ID": basic_info["bank_org_id"],
                    "GROUP_ID": "HZ"
                }
                sql_list.append(dict_to_sqls("GP_BM_ROLE_FUNC_REL", role_rel_info))
                # 按钮权限
                if function_info["pid"] == basic_info["module_name"].lower() + "_00" and function_info[
                    "path"] is not None:
                    for button_code in role_info["detail_buttons"]:
                        button_id = "BTN@" + function_info["id"] + "_" + button_code
                        role_rel_info = {
                            "ID": md5_str(str(role_info["id"]) + str(button_id)),
                            "ROLE_ID": role_info["id"],
                            "FUNCID": button_id,
                            "CORP_ID": basic_info["bank_code"],
                            "ORG_ID": basic_info["bank_org_id"],
                            "GROUP_ID": "HZ"
                        }
                        sql_list.append(dict_to_sqls("GP_BM_ROLE_FUNC_REL", role_rel_info))
            elif function_info["type"] == "按钮":
                button_id = "BTN@" + function_info["pid"] + "_" + function_info["id"]
                role_rel_info = {
                    "ID": md5_str(str(role_info["id"]) + str(button_id)),
                    "ROLE_ID": role_info["id"],
                    "FUNCID": button_id,
                    "CORP_ID": basic_info["bank_code"],
                    "ORG_ID": basic_info["bank_org_id"],
                    "GROUP_ID": "HZ"
                }
                sql_list.append(dict_to_sqls("GP_BM_ROLE_FUNC_REL", role_rel_info))


def gp_bm_business_line():
    sql_list.append(f"delete from GP_BM_BUSINESS_LINE")
    for business_line_info in business_line_list:
        data = {
            "DATA_ID": md5_str(str(business_line_info["id"])),
            "CORP_ID": basic_info["bank_code"],
            "ORG_ID": basic_info["bank_org_id"],
            "BUSINESS_LINE": business_line_info["id"],
            "BUSINESS_LINE_NAME": business_line_info["name"],
            "STATUS": "1",
        }
        add_default_cols(data, True)
        sql_list.append(dict_to_sqls("GP_BM_BUSINESS_LINE", data))


def gp_bm_business_line_role():
    sql_list.append(f"delete from GP_BM_BUSINESS_LINE_ROLE")
    for business_line_info in business_line_list:
        for role_name in business_line_info["roles"]:
            filtered = list(filter(lambda x: x["name"] == role_name, role_list))
            if len(filtered) == 0:
                raise Exception(
                    "业务线角色[{}-{}]初始化失败，角色名未找到对应角色".format(business_line_info["id"], role_name))
            role_info = filtered[0]
            data = {
                "DATA_ID": md5_str(str(business_line_info["id"]) + str(role_info["id"])),
                "CORP_ID": basic_info["bank_code"],
                "ORG_ID": basic_info["bank_org_id"],
                "BUSINESS_LINE": business_line_info["id"],
                "ROLE_ID": role_info["id"],
            }
            add_default_cols(data, True)
            sql_list.append(dict_to_sqls("GP_BM_BUSINESS_LINE_ROLE", data))


def gp_bm_res_tpl_inst():
    sql_list.append(f"delete from GP_BM_RES_TPL_INST")
    for business_line_info in business_line_list:
        business_line_id = business_line_info["id"]
        for privilege_info in business_line_info["privileges"]:
            report_code = privilege_info["report_code"]
            opr_org_ids = privilege_info["opr_org_ids"]
            data = {
                "DATA_ID": md5_str(str(business_line_id) + str(report_code)),
                "CORP_ID": basic_info["bank_code"],
                "ORG_ID": basic_info["bank_org_id"],
                "TPL_ID": report_code,
                "INST_ID": "TPL@" + business_line_id + "-" + report_code,
                "INST_NAME": business_line_id + "-" + report_code,
                "OPR_ID": business_line_id,
                "OPR_SCOPE": business_line_id,
                "IS_CUST": "0",
                "DATASET": upper_camel_case(basic_info["module_name"]) + report_code + "_record_ds",
                "OPR_ORGS": ",".join(opr_org_ids)
            }
            add_default_cols(data, True)
            sql_list.append(dict_to_sqls("GP_BM_RES_TPL_INST", data))


def gp_bm_res_tpl_field():
    sql_list.append(f"delete from GP_BM_RES_TPL_FIELD")
    for business_line_info in business_line_list:
        business_line_id = business_line_info["id"]
        for privilege_info in business_line_info["privileges"]:
            report_code = privilege_info["report_code"]
            for field in privilege_info["fields"]:
                data = {
                    "DATA_ID": md5_str(str(business_line_id) + str(report_code) + str(field["id"])),
                    "CORP_ID": basic_info["bank_code"],
                    "ORG_ID": basic_info["bank_org_id"],
                    "INST_ID": "TPL@" + business_line_id + "-" + report_code,
                    "FIELD_ID": field["id"],
                    "FIELD_NAME": field["name"],
                    "LIST_DISPLAY": "1",
                    "FORM_DISPLAY": "1",
                    "PERMISSION": "READ_WRITE"
                }
                add_default_cols(data, True)
                sql_list.append(dict_to_sqls("GP_BM_RES_TPL_FIELD", data))


def gp_bm_tlr_info():
    sql_list.append(f"delete from GP_BM_TLR_INFO")
    for user_info in user_list:
        data = {
            "DATA_ID": md5_str(str(user_info["id"])),
            "CORP_ID": basic_info["bank_code"],
            "ORG_ID": basic_info["bank_org_id"],
            "GROUP_ID": "HZ",
            "TLRNO": user_info["id"],
            "TLR_NAME": user_info["name"],
            "PASSWORD": "$2a$10$dwwKbUcFnqlaYTZy9BbE5uzWzdEgfgy4KPVXKqMCiUIFindJeLhRa",
            "PASSWD_ENC": "bcrypt",
            "BRCODE": user_info["org_id"],
            "BRNO": user_info["org_id"],
            "STATUS": "0",
            "ST": "4",
            "IS_LOCK": "0",
            "PSWD_ERR_CNT": 0,
            "TOT_PSWD_ERR_CNT": 0,
            "FLAG": "1",
            "CREDATE_DATE": curr_date_str,
            "LAST_UPD_OPR_ID": "admin",
            "LAST_UPD_TIME": curr_datetime_str,
            "LAST_PWD_CHG_TIME": curr_datetime_str,
            "EMAIL": user_info["email"],
            "GPMS_NEXTACTION": "21",
        }
        add_default_cols(data, True)
        sql_list.append(dict_to_sqls("GP_BM_TLR_INFO", data))


def gp_bm_tlr_role_rel():
    sql_list.append(f"delete from GP_BM_TLR_ROLE_REL")
    for user_info in user_list:
        for role_str in user_info["roles"]:
            role_name = role_str
            business_line_id = None
            if "," in role_name:
                role_name, business_line_id = role_str.split(",")
            filtered = list(filter(lambda x: x["name"] == role_name, role_list))
            if len(filtered) == 0:
                raise Exception("用户[{}]初始化失败，角色名[{}]未找到对应角色".format(user_info["id"], role_name))
            role_info = filtered[0]
            data = {
                "DATA_ID": md5_str(str(user_info["id"]) + str(role_info["id"]) + (
                    str(business_line_id) if business_line_id is not None else "")),
                "TLRNO": user_info["id"],
                "ROLE_ID": role_info["id"],
                "BUSINESS_LINE": business_line_id
            }
            add_default_cols(data, True)
            sql_list.append(dict_to_sqls("GP_BM_TLR_ROLE_REL", data))


def gp_bm_tlr_org_rel():
    sql_list.append(f"delete from GP_BM_TLR_ORG_REL")
    for user_info in user_list:
        for org_id in user_info["opr_orgs"]:
            data = {
                "DATA_ID": md5_str(str(user_info["id"]) + str(org_id)),
                "TLRNO": user_info["id"],
                "BRNO": org_id,
            }
            add_default_cols(data, True)
            sql_list.append(dict_to_sqls("GP_BM_TLR_ORG_REL", data))


def gp_bm_sys_param():
    for param_id, param_value in sys_param_config.items():
        sql_list.append(f"update GP_BM_SYS_PARAM set PARAM_VALUE = '{param_value}' where PARAM_ID = '{param_id}'")


def load_tpl_dict(wb):
    """
    加载字段
    """
    sheet_name = "字典"
    ws = wb[sheet_name]
    for row in ws.iter_rows(min_row=2):
        data_id = cv(row, 0)
        if data_id is None:
            continue
        dict_name = data_id
        dict_key = not_null(cv(row, 1), f"{sheet_name}-{data_id}-字典key")
        dict_value = not_null(cv(row, 2), f"{sheet_name}-{data_id}-字典value")
        if dict_name not in tpl_dict:
            tpl_dict[dict_name] = {}
        tpl_dict[dict_name][dict_value] = dict_key


def load_basic_info(wb):
    """
    加载基本信息
    """
    sheet_name = "基本信息"
    ws = wb[sheet_name]
    basic_info["tpl_version"] = not_null(ws["B1"].value, f"{sheet_name}-版本号")
    basic_info["bank_code"] = not_null(ws["B2"].value, f"{sheet_name}-机构编码")
    basic_info["bank_name"] = not_null(ws["B3"].value, f"{sheet_name}-机构名称")
    basic_info["bank_org_id"] = not_null(ws["B4"].value, f"{sheet_name}-总行机构号")
    basic_info["module_name"] = not_null(ws["B5"].value, f"{sheet_name}-产品模块")
    basic_info["sys_id"] = not_null(ws["B6"].value, f"{sheet_name}-产品ID")


def load_branches(wb):
    """
    加载机构信息
    """
    sheet_name = "机构"
    ws = wb[sheet_name]
    for row in ws.iter_rows(min_row=2):
        data_id = cv(row, 0)
        if data_id is None:
            continue
        branch_info = {
            "id": data_id,
            "name": not_null(cv(row, 1), f"{sheet_name}-{data_id}-行名"),
            "type": not_null(dv("机构类别", cv(row, 2)), f"{sheet_name}-{data_id}-机构类别"),
            "pid": not_null(cv(row, 3), f"{sheet_name}-{data_id}-上级机构代码"),
        }
        branch_list.append(branch_info)


def load_report_orgs(wb):
    """
    加载上报行
    """
    sheet_name = "上报行"

    branch_map = {}
    for branch in branch_list:
        branch_map[branch["id"]] = branch

    ws = wb[sheet_name]
    for row in ws.iter_rows(min_row=2):
        data_id = cv(row, 0)
        if data_id is None:
            continue
        report_org_info = {
            "id": data_id,
            "uni_credit_code": not_null(cv(row, 1), f"{sheet_name}-{data_id}-金融机构统一信用代码"),
            "name": not_null(cv(row, 2), f"{sheet_name}-{data_id}-机构名称"),
            "is_report": not_null(cv(row, 3), f"{sheet_name}-{data_id}-是否上报行"),
            "is_zh": "Y" if str(branch_map[data_id]["pid"]) == "0" else "N",
            "p_nbjgh": str(branch_map[data_id]["pid"]),
            "bank_org_code": not_null(cv(row, 4), f"{sheet_name}-{data_id}-12位银行机构代码"),
            "module_id": not_null(cv(row, 5), f"{sheet_name}-{data_id}-ModuleId"),
        }
        report_org_list.append(report_org_info)


def load_roles(wb):
    """
    加载角色
    """
    sheet_name = "角色"
    ws = wb[sheet_name]
    for row in ws.iter_rows(min_row=2):
        data_id = cv(row, 0)
        if data_id is None:
            continue
        role_info = {
            "id": data_id,
            "name": not_null(cv(row, 1), f"{sheet_name}-{data_id}-角色名"),
            "detail_buttons": lv(cv(row, 2)),
            "functions": []
        }
        role_list.append(role_info)


def load_business_lines(wb):
    """
    加载业务线
    """
    sheet_name = "业务线"
    ws = wb[sheet_name]
    for row in ws.iter_rows(min_row=2):
        data_id = cv(row, 0)
        if data_id is None:
            continue
        business_line_info = {
            "id": data_id,
            "name": not_null(cv(row, 1), f"{sheet_name}-{data_id}-业务线名称"),
            "roles": lv(cv(row, 2)),
            "privileges": []
        }
        business_line_list.append(business_line_info)


def load_users(wb):
    """
    加载用户
    """
    sheet_name = "用户"
    ws = wb[sheet_name]
    for row in ws.iter_rows(min_row=2):
        data_id = cv(row, 0)
        if data_id is None:
            continue
        role_str = not_null(cv(row, 4), f"{sheet_name}-{data_id}-用户角色")
        user_info = {
            "id": data_id,
            "name": not_null(cv(row, 1), f"{sheet_name}-{data_id}-柜员名"),
            "org_id": not_null(cv(row, 2), f"{sheet_name}-{data_id}-所属机构"),
            "email": not_null(cv(row, 3), f"{sheet_name}-{data_id}-邮箱"),
            "roles": role_str.split(";"),
            "opr_orgs": lv(cv(row, 5)),
        }
        user_list.append(user_info)


def load_functions(wb):
    """
    加载角色功能权限
    """
    sheet_name = "角色功能权限"
    ws = wb[sheet_name]

    role_column_index = ["G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T", "U", "V", "W", "X", "Y",
                         "Z", "AA", "AB", "AC", "AD", "AE", "AF", "AG",
                         "AH", "AI", "AJ", "AK", "AL", "AM", "AN", "AO", "AP", "AQ", "AR", "AS", "AT", "AU", "AV", "AW",
                         "AX",
                         "AY", "AZ",
                         "BA", "BB", "BC", "BD", "BE", "BF", "BG", "BH", "BI", "BJ", "BK", "BL", "BM", "BN", "BO", "BP",
                         "BQ", "BR", "BS", "BT", "BU", "BV", "BW", "BX",
                         "BY", "BZ",
                         "CA", "CB", "CC", "CD", "CE", "CF", "CG", "CH", "CI", "CJ", "CK", "CL", "CM", "CN", "CO", "CP",
                         "CQ", "CR", "CS", "CT", "CU", "CV", "CW", "CX",
                         "CY", "CZ",
                         "DA", "DB", "DC", "DD", "DE", "DF", "DG", "DH", "DI", "DJ", "DK", "DL", "DM", "DN", "DO", "DP",
                         "DQ", "DR", "DS", "DT", "DU", "DV", "DW", "DX",
                         "DY", "DZ", ]

    relation_role_name_list = [ws[f"{c}1"].value for c in role_column_index if ws[f"{c}1"].value is not None]

    for row in ws.iter_rows(min_row=2):
        data_id = cv(row, 1)
        if data_id is None:
            continue

        function_info = {
            "module": not_null(cv(row, 0), f"{sheet_name}-{data_id}-模块"),
            "id": data_id,
            "name": not_null(cv(row, 2), f"{sheet_name}-{data_id}-模块"),
            "type": not_null(cv(row, 3), f"{sheet_name}-{data_id}-模块"),
            "pid": not_null(cv(row, 4), f"{sheet_name}-{data_id}-模块"),
            "path": cv(row, 5)
        }
        for column_index in range(6, len(role_column_index)):
            if column_index - 6 >= len(relation_role_name_list):
                break
            role_name = relation_role_name_list[column_index - 6]
            cell_value = cv(row, column_index)
            if cell_value == "Y":
                for role_info in role_list:
                    if role_info["name"] == role_name:
                        role_info["functions"].append(function_info)
                        break
        function_list.append(function_info)


def load_object_privilege(wb):
    """
    加载业务线对象权限
    """
    sheet_name = "业务线对象权限"
    ws = wb[sheet_name]
    role_column_index = ["C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T", "U",
                         "V", "W", "X", "Y", "Z",
                         "AA", "AB", "AC", "AD", "AE", "AF", "AG", "AH", "AI", "AJ", "AK", "AL", "AM", "AN", "AO", "AP",
                         "AQ", "AR", "AS", "AT", "AU", "AV", "AW", "AX",
                         "AY", "AZ",
                         "BA", "BB", "BC", "BD", "BE", "BF", "BG", "BH", "BI", "BJ", "BK", "BL", "BM", "BN", "BO", "BP",
                         "BQ", "BR", "BS", "BT", "BU", "BV", "BW", "BX",
                         "BY", "BZ",
                         "CA", "CB", "CC", "CD", "CE", "CF", "CG", "CH", "CI", "CJ", "CK", "CL", "CM", "CN", "CO", "CP",
                         "CQ", "CR", "CS", "CT", "CU", "CV", "CW", "CX",
                         "CY", "CZ",
                         "DA", "DB", "DC", "DD", "DE", "DF", "DG", "DH", "DI", "DJ", "DK", "DL", "DM", "DN", "DO", "DP",
                         "DQ", "DR", "DS", "DT", "DU", "DV", "DW", "DX",
                         "DY", "DZ",
                         ]

    relation_business_line_id_list = [ws[f"{c}1"].value for c in role_column_index if ws[f"{c}1"].value is not None]
    for row in ws.iter_rows(min_row=2):
        report_code = cv(row, 0)
        if report_code is None:
            continue
        for column_index in range(2, 120):
            relation_list_index = column_index - 2
            if relation_list_index >= len(relation_business_line_id_list):
                break
            business_line_id = relation_business_line_id_list[relation_list_index]
            cell_value = cv(row, column_index)
            if cell_value is not None:
                for business_line_info in business_line_list:
                    if business_line_info["id"] == business_line_id:
                        if report_code not in report_field_list:
                            continue
                        business_line_info["privileges"].append({
                            "report_code": report_code,
                            "opr_org_ids": lv(cell_value),
                            "fields": report_field_list[report_code]
                        })
                        break


def load_reports(wb):
    """
    加载报表信息
    """
    sheet_name = "报表信息"
    ws = wb[sheet_name]
    for row in ws.iter_rows(min_row=2):
        report_code, report_name = cv(row, 0), cv(row, 1)
        if report_code is None:
            continue
        report_list.append({
            "id": report_code,
            "name": report_name
        })


def load_report_fields(wb):
    """
    加载报表字段信息
    """
    sheet_name = "报表字段信息"
    ws = wb[sheet_name]
    for row in ws.iter_rows(min_row=1):
        report_code, field_id, field_name = cv(row, 0), cv(row, 1), cv(row, 2)
        if report_code is None:
            continue
        if report_code not in report_field_list:
            report_field_list[report_code] = []
        report_field_list[report_code].append({
            "id": upper_score_to_lower_camel(field_id),
            "name": field_name
        })


def load_sys_param(wb):
    """
    加载系统初始化参数配置
    """
    sheet_name = "系统初始化参数配置"
    ws = wb[sheet_name]
    for row in ws:
        if row[1].value is None:
            continue
        sys_param_config[row[1].value] = not_null(row[3].value, f"{sheet_name}-{row[2].value}不能为空")


def not_null(value, name):
    if value is None or (type(value) == str and len(value) == 0):
        raise Exception(f"[{name}]不能为空")
    else:
        return str(value)


def cv(row, idx):
    """
    单元格值
    """
    return row[idx].value if len(row) > idx else None


def dv(dict_name, dict_value):
    """
    字典值转换
    """
    if dict_name not in tpl_dict or dict_value not in tpl_dict[dict_name]:
        return None
    else:
        return tpl_dict[dict_name][dict_value]


def lv(value):
    return [] if value is None else str(value).split(",")


def md5_str(src):
    return hashlib.md5(str(src).encode(encoding='UTF-8')).hexdigest()


def convert_value(value):
    if value is None:
        return "null"
    elif type(value) in [int, float]:
        return str(value)
    else:
        return "'" + value + "'"


def add_default_cols(cols, all=False):
    if all:
        cols['DATA_DATE'] = time.strftime("%Y%m%d", time.localtime())
        cols['DATA_SOURCE'] = 'O'
        cols['CHECK_FLAG'] = 'N'
        cols['DATA_VERSION'] = 0

    cols['NEXT_ACTION'] = '99'
    cols['DATA_STATUS'] = '04'
    cols['DATA_FLAG'] = '0'
    cols['DATA_CRT_USER'] = 'admin'
    return cols


def dict_to_sqls(table_name, data):
    columns_str = ",".join(["`%s`" % key if db_type == 'mysql' else key for key in data.keys()])
    values_str = ",".join([convert_value(value) for value in data.values()])
    return "insert into {}({}) values({})".format(table_name, columns_str, values_str)


def save_sql(output_file):
    with open(output_file, 'w', encoding="UTF-8") as f:
        f.writelines([sql + ("\ngo\n" if db_type == "sybase" else ";\n") for sql in sql_list])


def upper_score_to_lower_camel(source):
    upper_camel = "".join(map(lambda x: x.capitalize(), source.split("_")))
    return upper_camel[0].lower() + upper_camel[1:]


class Unbuffered(object):
    def __init__(self, stream):
        self.stream = stream

    def write(self, data):
        self.stream.write(data)
        self.stream.flush()

    def __getattr__(self, attr):
        return getattr(self.stream, attr)


if __name__ == '__main__':
    sys.stdout = Unbuffered(sys.stdout)
    sys.stderr = Unbuffered(sys.stderr)
    parser = argparse.ArgumentParser(description="RDMS init data generator.")
    parser.add_argument("-D", "--db", choices=["mysql", "oracle", 'sqlserver', 'sybase'], default="oracle",
                        help="db type:mysql oracle sqlserver sybase")
    logging.info("script version:" + script_version)
    args = parser.parse_args()
    logging.info('db type:{}'.format(args.db))
    db_type = args.db
    dist_file_path = WORKSPACE + 'init_' + time.strftime("%Y%m%d", time.localtime()) + '_' + db_type + '.sql'
    try:
        # 加载配置
        load_config(WORKSPACE + 'UCAM系统初始化.xlsx')
        # 生成sql
        gen_sql()
        # 写入文件
        save_sql(dist_file_path)
        logging.info("output file:" + dist_file_path)
    except Exception as e:
        logging.error("生成失败：", e)
        logging.info("生成失败：" + e.args[0])
